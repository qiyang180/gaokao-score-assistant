import argparse
import csv
import json
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_TEMPLATE = Path(__file__).resolve().parents[1].parent / "2026高考成绩汇总--理科.xlsx"
DEFAULT_SUBJECTS = [
    "总分",
    "语文",
    "数学",
    "外语",
    "物理",
    "历史",
    "化学",
    "生物",
    "思想政治",
    "地理",
]
EXTRA_HEADERS = ["查询状态", "截图路径", "错误原因", "查询时间"]
STUDENT_HEADER_ALIASES = {
    "班级": ("班级",),
    "学生姓名": ("姓名", "学生姓名", "考生姓名"),
    "身份证号码": ("身份证号", "身份证号码", "证件号", "证件号码"),
    "考生号": ("考生号", "准考证号"),
    "报名序号": ("报名序号", "报名号"),
}
SCORE_HEADER_ALIASES = {
    "英语": ("英语", "外语"),
    "生物": ("生物", "生物学"),
    "生物/政治/地理": ("生物/政治/地理", "生物", "生物学", "思想政治", "政治", "地理"),
}


def load_results(path: Path):
    if not path.exists():
        raise FileNotFoundError(f"results file not found: {path}")

    results = []
    for line_no, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        if not line.strip():
            continue
        try:
            results.append(json.loads(line))
        except json.JSONDecodeError as exc:
            raise ValueError(f"invalid JSON at line {line_no}: {exc}") from exc
    return results


def load_students(path: Path):
    if not path.exists():
        return []

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return [
            {str(key or "").strip(): str(value or "").strip() for key, value in row.items()}
            for row in csv.DictReader(f)
        ]


def collect_subjects(results):
    subjects = list(DEFAULT_SUBJECTS)
    seen = set(subjects)
    for item in results:
        for key in (item.get("scores") or {}).keys():
            if key not in seen:
                seen.add(key)
                subjects.append(key)
    return subjects


def get_first(mapping, aliases):
    for key in aliases:
        value = mapping.get(key, "")
        if value not in (None, ""):
            return value
    return ""


def result_student_name(item):
    return ((item.get("student") or {}).get("name") or "").strip()


def match_students(results, students):
    by_name = {}
    for student in students:
        name = get_first(student, STUDENT_HEADER_ALIASES["学生姓名"]).strip()
        if name:
            by_name.setdefault(name, []).append(student)

    matched = []
    for item in results:
        name = result_student_name(item)
        candidates = by_name.get(name) or []
        matched.append(candidates.pop(0) if candidates else {})
    return matched


def autosize_columns(ws):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)


def copy_cell_style(source, target):
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)


def load_template(template_path: Path):
    if template_path.exists():
        return load_workbook(template_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "成绩汇总"
    ws.append(["姓名", *collect_subjects([]), *EXTRA_HEADERS])
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    return wb


def ensure_headers(ws, headers):
    header_map = {}
    for col in range(1, ws.max_column + 1):
        value = str(ws.cell(1, col).value or "").strip()
        if value:
            header_map[value] = col

    style_source = ws.cell(1, ws.max_column if ws.max_column else 1)
    next_col = ws.max_column + 1
    for header in headers:
        if header in header_map:
            continue
        cell = ws.cell(1, next_col, header)
        copy_cell_style(style_source, cell)
        header_map[header] = next_col
        next_col += 1
    return header_map


def clear_output_area(ws, header_map, data_rows):
    max_row = max(ws.max_row, data_rows + 1)
    for row in range(2, max_row + 1):
        for col in header_map.values():
            ws.cell(row, col).value = None


def value_for_header(header, item, student):
    scores = item.get("scores") or {}

    if header in STUDENT_HEADER_ALIASES:
        return get_first(student, STUDENT_HEADER_ALIASES[header])
    if header == "姓名":
        return get_first(student, STUDENT_HEADER_ALIASES["学生姓名"]) or result_student_name(item)
    if header == "查询状态":
        return item.get("status", "")
    if header == "截图路径":
        return item.get("screenshotPath", "")
    if header == "错误原因":
        return item.get("error", "")
    if header == "查询时间":
        return item.get("queriedAt", "")
    if header in SCORE_HEADER_ALIASES:
        return get_first(scores, SCORE_HEADER_ALIASES[header])
    return scores.get(header, "")


def build_workbook(results, out_path: Path, students=None, template_path=DEFAULT_TEMPLATE):
    subjects = collect_subjects(results)
    wb = load_template(Path(template_path))
    ws = wb.active

    template_headers = [
        str(ws.cell(1, col).value or "").strip()
        for col in range(1, ws.max_column + 1)
        if str(ws.cell(1, col).value or "").strip()
    ]
    headers = template_headers or ["姓名", *subjects]
    header_map = ensure_headers(ws, [*headers, *EXTRA_HEADERS])
    clear_output_area(ws, header_map, len(results))

    matched_students = match_students(results, students or [])
    style_row = 2 if ws.max_row >= 2 else 1
    for offset, (item, student) in enumerate(zip(results, matched_students), start=2):
        for header, col in header_map.items():
            cell = ws.cell(offset, col, value_for_header(header, item, student))
            copy_cell_style(ws.cell(style_row, col), cell)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(description="Build Excel summary from query JSONL results.")
    parser.add_argument("--results", default="output/results.jsonl")
    parser.add_argument("--students", default="work/students.csv")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE))
    parser.add_argument("--out", default="output/成绩汇总.xlsx")
    args = parser.parse_args()

    results = load_results(Path(args.results))
    students = load_students(Path(args.students))
    build_workbook(results, Path(args.out), students=students, template_path=Path(args.template))
    print(f"summary written: {args.out}")


if __name__ == "__main__":
    main()
