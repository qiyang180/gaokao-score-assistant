import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_NAME = "姓名"
ID_COLUMNS = ("身份证号", "准考证号", "考生号", "报名序号")
OUTPUT_COLUMNS = ("班级", "姓名", "身份证号", "准考证号", "考生号", "报名序号")
FIELD_ALIASES = {
    "班级": ("班级", "行政班", "班别"),
    "姓名": ("姓名", "考生姓名", "学生姓名"),
    "身份证号": ("身份证号", "身份证号码", "证件号", "证件号码", "居民身份证号"),
    "准考证号": ("准考证号", "准考证号码"),
    "考生号": ("考生号", "考生号码"),
    "报名序号": ("报名序号", "报名号"),
}


def normalize_header(value):
    return str(value or "").strip().replace(" ", "")


def trim_trailing_empty(values):
    values = list(values)
    while values and values[-1] in (None, ""):
        values.pop()
    return values


def looks_like_header(headers):
    header_set = {header for header in headers if header}
    if not any(alias in header_set for alias in FIELD_ALIASES["姓名"]):
        return False
    return any(
        alias in header_set
        for output_col in ID_COLUMNS
        for alias in FIELD_ALIASES[output_col]
    )


def normalize_record(record):
    normalized = {"__row": record.get("__row")}
    for output_col, aliases in FIELD_ALIASES.items():
        normalized[output_col] = ""
        for alias in aliases:
            value = record.get(alias, "")
            if value:
                normalized[output_col] = value
                break
    return normalized


def read_xlsx(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [trim_trailing_empty(row) for row in ws.iter_rows(values_only=True)]
    if not rows:
        return []

    header_index = None
    headers = []
    for index, row in enumerate(rows[:20]):
        candidate = [normalize_header(value) for value in row]
        if looks_like_header(candidate):
            header_index = index
            headers = candidate
            break

    if header_index is None:
        headers = [normalize_header(value) for value in rows[0]]
        header_index = 0

    records = []
    for row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        record = {}
        for col_index, header in enumerate(headers):
            if not header:
                continue
            value = row[col_index] if col_index < len(row) else ""
            record[header] = "" if value is None else str(value).strip()
        record["__row"] = row_index
        if any(record.get(col, "") for col in headers):
            records.append(record)
    return records


def read_csv(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        records = []
        for row_index, row in enumerate(reader, start=2):
            record = {normalize_header(k): str(v or "").strip() for k, v in row.items()}
            record["__row"] = row_index
            if any(value for key, value in record.items() if key != "__row"):
                records.append(record)
        return records


def validate(records):
    if not records:
        raise ValueError("学生表为空")

    missing = []
    for record in records:
        if not record.get(REQUIRED_NAME):
            missing.append(f"第 {record.get('__row')} 行缺少姓名")
        if not any(record.get(col) for col in ID_COLUMNS):
            missing.append(f"第 {record.get('__row')} 行缺少身份证号/准考证号/考生号/报名序号")

    if missing:
        raise ValueError("\n".join(missing))


def write_csv(records, out_path: Path):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        for record in records:
            writer.writerow({col: record.get(col, "") for col in OUTPUT_COLUMNS})


def main():
    parser = argparse.ArgumentParser(description="Normalize student xlsx/csv into the query CSV format.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--out", default="work/students.csv")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        raise FileNotFoundError(f"input file not found: {input_path}")

    if input_path.suffix.lower() == ".xlsx":
        records = read_xlsx(input_path)
    elif input_path.suffix.lower() == ".csv":
        records = read_csv(input_path)
    else:
        raise ValueError("只支持 .xlsx 或 .csv 学生表")

    records = [normalize_record(record) for record in records]
    validate(records)
    write_csv(records, Path(args.out))
    print(f"normalized {len(records)} students -> {args.out}")


if __name__ == "__main__":
    main()
